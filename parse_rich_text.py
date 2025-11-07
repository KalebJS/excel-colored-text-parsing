#!/usr/bin/env python3
"""
Parse Excel cells with rich text formatting and extract color information.
"""
from colorsys import hls_to_rgb, rgb_to_hls
from datetime import date, datetime, time, timedelta
from decimal import Decimal
from typing import Any, Literal

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell as OpenpyxlCell
from openpyxl.cell.rich_text import CellRichText
from openpyxl.styles.fonts import Font
from openpyxl.styles.proxy import StyleProxy
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.xml.functions import QName, fromstring
from pydantic import BaseModel

# Type alias for cell values based on openpyxl's internal types
CellValue = (
    float
    | Decimal
    | str
    | CellRichText
    | datetime
    | date
    | time
    | timedelta
    | Literal[True]
    | Any  # For formula types and other edge cases
    | None
)


class Segment(BaseModel):
    """Represents a text segment with RGB color information."""
    r: int
    g: int
    b: int
    text: str
    is_default_color: bool = False  # True if using fallback black color (no rich text or cell color)
    
    @property
    def is_black(self) -> bool:
        """Check if color is black (with some leeway)."""
        # Allow up to 30 in each channel to be considered "black"
        return self.r <= 30 and self.g <= 30 and self.b <= 30
    
    @property
    def is_red(self) -> bool:
        """Check if color is red (with some leeway)."""
        # Red should be high (>200), green and blue should be low (<80)
        return self.r > 200 and self.g < 80 and self.b < 80
    
    @property
    def is_blue(self) -> bool:
        """Check if color is blue (with some leeway)."""
        # Blue should be high (>200), red and green should be low (<80)
        return self.r < 80 and self.g < 80 and self.b > 200


class Cell(BaseModel):
    """Represents an Excel cell with rich text segments."""
    cell_number: str
    color_groups: list[Segment]


# Standard indexed colors (indices 0-63)
INDEXED_COLORS = [
    '00000000', '00FFFFFF', '00FF0000', '0000FF00', '000000FF',
    '00FFFF00', '00FF00FF', '0000FFFF', '00000000', '00FFFFFF',
    '00FF0000', '0000FF00', '000000FF', '00FFFF00', '00FF00FF',
    '0000FFFF', '00800000', '00008000', '00000080', '00808000',
    '00800080', '00008080', '00C0C0C0', '00808080', '009999FF',
    '00993366', '00FFFFCC', '00CCFFFF', '00660066', '00FF8080',
    '000066CC', '00CCCCFF', '00000080', '00FF00FF', '00FFFF00',
    '0000FFFF', '00800080', '00800000', '00008080', '000000FF',
    '0000CCFF', '00CCFFFF', '00CCFFCC', '00FFFF99', '0099CCFF',
    '00FF99CC', '00CC99FF', '00FFCC99', '003366FF', '0033CCCC',
    '0099CC00', '00FFCC00', '00FF9900', '00FF6600', '00666699',
    '00969696', '00003366', '00339966', '00003300', '00333300',
    '00993300', '00993366', '00333399', '00333333'
]

# Constants for color conversion
RGBMAX = 0xff  # 255
HLSMAX = 240  # MS Excel's HLS is base 240


def hex_to_rgb(hex_color: str) -> tuple[int, int, int]:
    """Convert hex color string to RGB tuple."""
    # Remove any leading '#' if present
    hex_color = hex_color.lstrip('#')
    
    # Ensure we have 6 characters
    if len(hex_color) != 6:
        return (0, 0, 0)
    
    r = int(hex_color[0:2], 16)
    g = int(hex_color[2:4], 16)
    b = int(hex_color[4:6], 16)
    
    return (r, g, b)


def rgb_to_ms_hls(red: int | str, green: int | None = None, blue: int | None = None) -> tuple[int, int, int]:
    """
    Converts RGB values (0-255) or hex string to MS Excel HLS format (base 240).
    Based on: https://gist.github.com/Mike-Honey/b36e651e9a7f1d2e1d60ce1c63b9b633
    """
    if green is None:
        if isinstance(red, str):
            if len(red) > 6:
                red = red[-6:]  # Ignore alpha
            blue_val = int(red[4:], 16) / RGBMAX
            green_val = int(red[2:4], 16) / RGBMAX
            red_val = int(red[0:2], 16) / RGBMAX
        else:
            raise ValueError("Invalid RGB input")
    else:
        red_val = red / RGBMAX
        green_val = green / RGBMAX
        blue_val = blue / RGBMAX
    
    h, l, s = rgb_to_hls(red_val, green_val, blue_val)
    return (int(round(h * HLSMAX)), int(round(l * HLSMAX)), int(round(s * HLSMAX)))


def ms_hls_to_rgb(hue: int, lightness: int, saturation: int) -> tuple[float, float, float]:
    """Converts MS Excel HLS (base 240) to RGB (0-1 range)."""
    return hls_to_rgb(hue / HLSMAX, lightness / HLSMAX, saturation / HLSMAX)


def rgb_to_hex(red: float, green: float, blue: float) -> str:
    """Converts RGB values (0-1 range) to hex string."""
    return ('%02x%02x%02x' % (
        int(round(red * RGBMAX)),
        int(round(green * RGBMAX)),
        int(round(blue * RGBMAX))
    )).upper()


def tint_luminance(tint: float, lum: int) -> int:
    """Apply tint to luminance value."""
    if tint < 0:
        return int(round(lum * (1.0 + tint)))
    else:
        return int(round(lum * (1.0 - tint) + (HLSMAX - HLSMAX * (1.0 - tint))))


def get_theme_colors(wb: Workbook) -> list[str]:
    """
    Extract theme colors from workbook.
    Based on: https://gist.github.com/Mike-Honey/b36e651e9a7f1d2e1d60ce1c63b9b633
    """
    if not hasattr(wb, 'loaded_theme') or wb.loaded_theme is None:
        # Return default Office theme colors if no theme is loaded
        return [
            'FFFFFF', '000000', 'EEECE1', '1F497D', '4F81BD', 'C0504D',
            '9BBB59', '8064A2', '4BACC6', 'F79646'
        ]
    
    xlmns = 'http://schemas.openxmlformats.org/drawingml/2006/main'
    root = fromstring(wb.loaded_theme)
    theme_el = root.find(QName(xlmns, 'themeElements').text)
    color_schemes = theme_el.findall(QName(xlmns, 'clrScheme').text)
    first_color_scheme = color_schemes[0]
    
    colors = []
    for c in ['lt1', 'dk1', 'lt2', 'dk2', 'accent1', 'accent2', 'accent3', 'accent4', 'accent5', 'accent6']:
        accent = first_color_scheme.find(QName(xlmns, c).text)
        color_elements = list(accent)
        if color_elements and 'window' in color_elements[0].attrib.get('val', ''):
            colors.append(color_elements[0].attrib.get('lastClr', '000000'))
        elif color_elements:
            colors.append(color_elements[0].attrib.get('val', '000000'))
        else:
            colors.append('000000')
    
    return colors


def theme_and_tint_to_rgb(wb: Workbook, theme: int, tint: float) -> str:
    """Convert theme color with tint to RGB hex."""
    theme_colors = get_theme_colors(wb)
    if theme >= len(theme_colors):
        return '000000'
    
    rgb_hex = theme_colors[theme]
    h, l, s = rgb_to_ms_hls(rgb_hex)
    return rgb_to_hex(*ms_hls_to_rgb(h, tint_luminance(tint, l), s))


def get_color_hex(
    font: Font | StyleProxy | None, 
    wb: Workbook,
    default_color: str = '000000'
) -> tuple[str, bool]:
    """
    Extract color hex from font object, resolving theme and indexed colors.
    
    Args:
        font: Font or StyleProxy object
        wb: Workbook (needed for theme color resolution)
        default_color: Fallback color if no color is set
    
    Returns:
        Tuple of (color_hex, is_default) where is_default is True if using fallback color
    """
    if not font or not font.color:
        return (default_color, True)
    
    color_obj = font.color
    
    # RGB format (ARGB - first 2 chars are alpha channel)
    if color_obj.rgb:
        rgb_value = str(color_obj.rgb)
        # Strip alpha channel (first 2 chars) to get RGB
        color_hex = rgb_value[2:] if len(rgb_value) > 6 else rgb_value
        return (color_hex, False)
    
    # Theme color with optional tint
    if color_obj.theme is not None:
        tint = color_obj.tint if color_obj.tint else 0.0
        color_hex = theme_and_tint_to_rgb(wb, color_obj.theme, tint)
        return (color_hex, False)
    
    # Indexed color
    if color_obj.indexed is not None:
        idx = color_obj.indexed
        if 0 <= idx < len(INDEXED_COLORS):
            # Strip alpha channel from indexed color
            color_hex = INDEXED_COLORS[idx][2:]
            return (color_hex, False)
    
    # No color information available
    return (default_color, True)


def parse_rich_text_cell(
    cell_value: CellValue, 
    cell_ref: str,
    wb: Workbook,
    cell_font: Font | StyleProxy | None = None
) -> Cell:
    """
    Parse a cell with rich text and return a Cell object with color segments.
    
    Args:
        cell_value: The value of the cell (can be rich text or plain text)
        cell_ref: The cell reference (e.g., "A1")
        wb: Workbook (needed for theme color resolution)
        cell_font: The cell's base font (used to get default color for unformatted segments)
    """
    segments: list[Segment] = []
    
    # Get the default color from the cell's font, or use black as fallback
    default_color, cell_uses_default = get_color_hex(cell_font, wb, '000000')
    
    if cell_value is None:
        return Cell(cell_number=cell_ref, color_groups=[])
    
    if not isinstance(cell_value, CellRichText):
        # Plain text cell - no rich formatting, use cell's font color
        r, g, b = hex_to_rgb(default_color)
        segment = Segment(r=r, g=g, b=b, text=str(cell_value), is_default_color=cell_uses_default)
        return Cell(cell_number=cell_ref, color_groups=[segment])
    
    # Process rich text segments
    for item in cell_value:
        if isinstance(item, str):
            # Unformatted text segment - use cell's base font color
            r, g, b = hex_to_rgb(default_color)
            segments.append(Segment(r=r, g=g, b=b, text=item, is_default_color=cell_uses_default))
        else:
            # Formatted text segment with font styling (TextBlock)
            text = item.text
            color_hex, is_default = get_color_hex(item.font, wb, default_color)
            r, g, b = hex_to_rgb(color_hex)
            segments.append(Segment(r=r, g=g, b=b, text=text, is_default_color=is_default))
    
    return Cell(cell_number=cell_ref, color_groups=segments)


def main() -> list[Cell]:
    # Load workbook with rich text support enabled
    print("Loading Book.xlsx with rich text support...")
    wb = load_workbook('Book.xlsx', rich_text=True)
    ws = wb.active
    
    if not isinstance(ws, Worksheet):
        raise ValueError("Active sheet is not a valid worksheet")
    
    print(f"\nWorksheet: {ws.title}")
    print("=" * 80)
    
    cells: list[Cell] = []
    
    # Iterate through all rows and cells
    for row in ws.iter_rows():
        for cell in row:
            if not isinstance(cell, OpenpyxlCell):
                continue
            if cell.value:
                # Parse the cell into a Cell object
                cell_ref = f"{cell.column_letter}{cell.row}"
                # Pass the workbook and cell's font to get the correct color
                cell_obj = parse_rich_text_cell(cell.value, cell_ref, wb, cell.font)
                cells.append(cell_obj)
                
                # Print cell information
                print(f"\nCell {cell_obj.cell_number}:")
                print(f"  Total segments: {len(cell_obj.color_groups)}")
                
                for i, segment in enumerate(cell_obj.color_groups, 1):
                    print(f"\n  Segment {i}:")
                    print(f"    RGB: ({segment.r}, {segment.g}, {segment.b})")
                    print(f"    Text: {segment.text[:50]}{'...' if len(segment.text) > 50 else ''}")
                    print(f"    is_default_color: {segment.is_default_color}")
                    print(f"    is_black: {segment.is_black}")
                    print(f"    is_red: {segment.is_red}")
                    print(f"    is_blue: {segment.is_blue}")
                
                print()
    
    print("=" * 80)
    print(f"\nTotal cells processed: {len(cells)}")
    
    return cells


if __name__ == "__main__":
    cells = main()
